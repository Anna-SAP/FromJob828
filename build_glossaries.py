#!/usr/bin/env python3
"""
Parse TBX and Glossary xlsx files and produce structured JSON per language.

Data sources:
  _TBX/    — authoritative terminology (sheet: "Terms")
  _Glossaries/ — supplementary glossaries (latest file per language)

Output:
  output/glossaries/<lang-code>.json   — one file per language
  output/glossaries/_summary.json      — overview of all languages
"""

import json
import os
import re
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl
import pandas as pd

ROOT = Path(__file__).resolve().parent
TBX_DIR = ROOT / "_TBX"
GLOSS_DIR = ROOT / "_Glossaries"
OUT_DIR = ROOT / "output" / "glossaries"

# ---------------------------------------------------------------------------
# Language-code normalisation helpers
# ---------------------------------------------------------------------------

# Map from TBX filename lang codes (lowercase) to canonical codes used in output
LANG_NORMALISE = {
    "de-de": "de-de",
    "es-419": "es-419",
    "es-es": "es-es",
    "fi-fi": "fi-fi",
    "fr-ca": "fr-ca",
    "fr-fr": "fr-fr",
    "it-it": "it-it",
    "ja-jp": "ja-jp",
    "ko-kr": "ko-kr",
    "nl-nl": "nl-nl",
    "pl-pl": "pl-pl",
    "pt-br": "pt-br",
    "pt-pt": "pt-pt",
    "ro-ro": "ro-ro",
    "sv-se": "sv-se",
    "zh-cn": "zh-cn",
    "zh-tw": "zh-tw",
}


def norm_lang(raw: str) -> str:
    """Return canonical lower-case lang code."""
    raw = raw.strip().lower().replace("_", "-")
    return LANG_NORMALISE.get(raw, raw)


# ---------------------------------------------------------------------------
# TBX parsing
# ---------------------------------------------------------------------------

def extract_lang_from_tbx_filename(fname: str) -> str:
    """Extract language code from e.g. terms_en_zh-cn_2026-01-21T05-04-23.xlsx"""
    m = re.match(r"terms_en_([a-z]{2}-[a-z0-9]+)_", fname, re.I)
    if m:
        return norm_lang(m.group(1))
    return ""


def parse_tbx(filepath: Path) -> list[dict]:
    """Parse one TBX xlsx and return list of term dicts."""
    wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
    ws = wb["Terms"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []

    header = [str(c).strip().lower() if c else "" for c in rows[0]]
    # Expected: concept id, source term (en), target term (...), definition,
    #           part of speech, context, status, last mod by, last mod date, notes
    col = {}
    for i, h in enumerate(header):
        if "concept" in h and "id" in h:
            col["concept_id"] = i
        elif "source" in h and "term" in h:
            col["source_term"] = i
        elif "target" in h and "term" in h:
            col["target_term"] = i
        elif h == "definition":
            col["definition"] = i
        elif "part" in h and "speech" in h:
            col["part_of_speech"] = i
        elif h == "context":
            col["context"] = i
        elif h == "status":
            col["status"] = i
        elif "last" in h and "date" in h:
            col["last_mod_date"] = i
        elif h == "notes":
            col["notes"] = i

    terms = []
    for row in rows[1:]:
        def cell(key):
            idx = col.get(key)
            if idx is None or idx >= len(row):
                return None
            v = row[idx]
            if v is None:
                return None
            if isinstance(v, (date, datetime)):
                return v.isoformat()[:10]
            return str(v).strip() if v else None

        src = cell("source_term")
        tgt = cell("target_term")
        if not src or not tgt:
            continue

        status_raw = cell("status")
        status = None
        if status_raw:
            su = status_raw.upper()
            if "VALID" in su:
                status = "VALID"
            elif "LEGACY" in su or "DEPREC" in su:
                status = "LEGACY"
            else:
                status = status_raw

        terms.append({
            "source_term": src,
            "target_term": tgt,
            "part_of_speech": cell("part_of_speech") or None,
            "definition": cell("definition") or None,
            "context": cell("context") or None,
            "status": status,
            "notes": cell("notes") or None,
            "last_mod_date": cell("last_mod_date") or None,
        })
    return terms


# ---------------------------------------------------------------------------
# Glossary language detection & date extraction
# ---------------------------------------------------------------------------

# Mapping from glossary filename patterns to language codes.
# Ordered so more specific patterns match first.
GLOSS_LANG_MAP = [
    # RingCX glossaries
    (r"RingCX_glossary_fr-CA", "fr-ca"),
    (r"RingCX_glossary_es-419", "es-419"),
    (r"RingCX_glossary_French", "fr-fr"),
    (r"RingCX_glossary_German", "de-de"),
    (r"RingCX_glossary_Italian", "it-it"),
    (r"RingCX_glossary_Spanish", "es-es"),
    # Engage Digital — FR-FR product glossary
    (r"Engage Digital Glossary EN-FR", "fr-fr"),
    # RC Glossary standard naming
    (r"RC[_ ]Glossary[_ ]*_?ZH-CN", "zh-cn"),
    (r"RC[_ ]Glossary[_ ]*_?ZH-TW", "zh-tw"),
    (r"RC[_ ]Glossary[_ ]*_?DE-DE", "de-de"),
    (r"RC[_ ]Glossary[_ ]*_?EN-UK", "en-uk"),
    (r"RC[_ ]Glossary[_ ]*_?ES-ES", "es-es"),
    (r"RC[_ ]Glossary[_ ]*_?ES-XL", "es-419"),
    (r"RC[_ ]Glossary[_ ]*_?fi-FI", "fi-fi"),
    (r"RC[_ ]Glossary[_ ]*_?FR-CA", "fr-ca"),
    (r"RC[_ ]Glossary[_ ]*_?FR-FR", "fr-fr"),
    (r"RC[_ ]Glossary[_ ]*_?IT-IT", "it-it"),
    (r"RC[_ ]Glossary[_ ]*_?JA-JP", "ja-jp"),
    (r"RC[_ ]Glossary[_ ]*_?KO-KR", "ko-kr"),
    (r"RC[_ ]Glossary[_ ]*_?NL[_ ]NL", "nl-nl"),
    (r"RC[_ ]Glossary[_ ]*_?PL-PL", "pl-pl"),
    (r"RC[_ ]Glossary[_ ]*_?PT-BR", "pt-br"),
    (r"RC[_ ]Glossary[_ ]*_?PT-PT", "pt-pt"),
    (r"RC[_ ]Glossary[_ ]*_?ro-RO", "ro-ro"),
    (r"RC[_ ]Glossary[_ ]*_?SV-SE", "sv-se"),
    # Polish special case
    (r"RC glossary.*Polish", "pl-pl"),
]


def detect_gloss_lang(fname: str) -> str | None:
    """Detect language code from glossary filename."""
    for pattern, lang in GLOSS_LANG_MAP:
        if re.search(pattern, fname, re.I):
            return lang
    return None


def extract_date_from_filename(fname: str) -> str:
    """Extract the latest date-like component from a filename for sorting.
    Returns a string that sorts chronologically (YYYYMMDD).
    """
    stem = Path(fname).stem
    # Try YYYYMMDD pattern (e.g., 20240709, 20240428)
    m = re.findall(r"(20\d{6})", stem)
    if m:
        return max(m)
    # Try YYYY_MM_DD or MM_DD_YYYY patterns
    m = re.findall(r"(\d{4})[_-](\d{2})[_-](\d{2})", stem)
    if m:
        candidates = []
        for g in m:
            y, a, b = g
            if int(y) >= 2000:
                candidates.append(f"{y}{a}{b}")
        if candidates:
            return max(candidates)
    # Try MM_DD_YYYY
    m = re.findall(r"(\d{2})[_-](\d{2})[_-](\d{4})", stem)
    if m:
        candidates = []
        for g in m:
            a, b, y = g
            if int(y) >= 2000:
                candidates.append(f"{y}{a}{b}")
        if candidates:
            return max(candidates)
    # Try YYMMDD (e.g., 240701)
    m = re.findall(r"(\d{6})", stem)
    if m:
        for candidate in m:
            yy = int(candidate[:2])
            if 20 <= yy <= 30:
                return f"20{candidate}"
    # Fallback: "2020" in name like "RC glossary 2020 Polish"
    m = re.findall(r"(20\d{2})", stem)
    if m:
        return max(m) + "0101"
    return "00000000"


def pick_latest_glossary_per_lang(gloss_dir: Path) -> dict[str, Path]:
    """Return {lang_code: filepath} using the latest file per language."""
    candidates: dict[str, list[tuple[str, Path]]] = {}
    for f in gloss_dir.iterdir():
        if not f.suffix.lower() in (".xlsx", ".xls"):
            continue
        if f.name.startswith("~$"):
            continue
        lang = detect_gloss_lang(f.name)
        if lang is None:
            print(f"  [WARN] Cannot detect language for glossary: {f.name}", file=sys.stderr)
            continue
        # Skip en-uk — we don't have a TBX for it and it's EN→EN
        if lang == "en-uk":
            continue
        d = extract_date_from_filename(f.name)
        candidates.setdefault(lang, []).append((d, f))

    result = {}
    for lang, files in candidates.items():
        files.sort(key=lambda x: x[0], reverse=True)
        result[lang] = files[0][1]
    return result


# ---------------------------------------------------------------------------
# Glossary parsing — flexible column detection
# ---------------------------------------------------------------------------

def _header_lower(row) -> list[str]:
    return [str(c).strip().lower() if c else "" for c in row]


def detect_glossary_columns(header: list[str], filepath: Path) -> dict:
    """Detect column indices for source, target, definition, notes from header row."""
    h = _header_lower(header)
    cols: dict[str, int | None] = {
        "source": None,
        "target": None,
        "definition": None,
        "notes": None,
    }
    fname = filepath.name.lower()

    # --- Source term column (always column 0 for RC/RingCX glossaries) ---
    # Special case: Engage Digital has Keys / English / French
    if "engage digital" in fname:
        cols["source"] = 1  # "English" column
        cols["target"] = 2  # "French" column
        # definition/notes not present in this format
        for i, hv in enumerate(h):
            if "definition" in hv:
                cols["definition"] = i
            if "notes" in hv and cols["notes"] is None:
                cols["notes"] = i
        return cols

    # Standard: first column is English source
    for i, hv in enumerate(h):
        if any(kw in hv for kw in ["english", "en_us", "en-us", "en_uk"]):
            cols["source"] = i
            break
    if cols["source"] is None:
        cols["source"] = 0  # fallback

    # --- Target term column ---
    # For files with multiple candidate columns (ZH-CN, ZH-TW, ro-RO),
    # pick the approved/final column; fall back to the column right after source.
    target_candidates = []
    for i, hv in enumerate(h):
        if i == cols["source"]:
            continue
        # Skip comment/meta columns
        if any(skip in hv for skip in ["comment", "joncers", "linguist", "project",
                                        "file name", "date", "reference", "background"]):
            continue
        # Prefer "cleared up & approved", "final", or explicit lang code headers
        if "approved" in hv or "cleared" in hv:
            target_candidates.insert(0, (i, 100))  # highest priority
        elif "final" in hv:
            target_candidates.append((i, 90))
        elif hv and not any(kw in hv for kw in ["definition", "notes", "english",
                                                  "en_us", "en-us", "en_uk",
                                                  "secondary", "original"]):
            # Check if this looks like a language column
            lang_patterns = [
                "german", "french", "spanish", "italian", "portuguese", "dutch",
                "polish", "swedish", "finnish", "japanese", "korean", "chinese",
                "target", "ro-ro", "zh-cn", "zh-tw", "de-de", "fr-fr", "fr-ca",
                "es-es", "es-419", "it-it", "ja-jp", "ko-kr", "nl-nl", "pl-pl",
                "pt-br", "pt-pt", "sv-se", "fi-fi",
                "de_de", "fr_fr", "fr_ca", "es_es", "it_it",
            ]
            for lp in lang_patterns:
                if lp in hv:
                    target_candidates.append((i, 80))
                    break
            else:
                # Column right after source that isn't recognized
                if i == cols["source"] + 1:
                    target_candidates.append((i, 50))

    if target_candidates:
        target_candidates.sort(key=lambda x: -x[1])
        cols["target"] = target_candidates[0][0]
    else:
        cols["target"] = cols["source"] + 1  # fallback

    # --- Definition ---
    for i, hv in enumerate(h):
        if "definition" in hv:
            cols["definition"] = i
            break

    # --- Notes ---
    for i, hv in enumerate(h):
        if i == cols["definition"]:
            continue
        if "notes" in hv and "comment" not in hv:
            cols["notes"] = i
            break

    return cols


def parse_glossary(filepath: Path) -> list[dict]:
    """Parse one glossary xlsx and return list of term dicts."""
    wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []

    cols = detect_glossary_columns(list(rows[0]), filepath)
    src_i = cols["source"]
    tgt_i = cols["target"]
    def_i = cols["definition"]
    notes_i = cols["notes"]

    terms = []
    for row in rows[1:]:
        def cell(idx):
            if idx is None or idx >= len(row):
                return None
            v = row[idx]
            if v is None:
                return None
            s = str(v).strip()
            # Skip formula remnants
            if s.startswith("="):
                return None
            return s if s else None

        src = cell(src_i)
        tgt = cell(tgt_i)
        if not src or not tgt:
            continue

        terms.append({
            "source_term": src,
            "target_term": tgt,
            "part_of_speech": None,
            "definition": cell(def_i) or None,
            "context": None,
            "status": None,
            "notes": cell(notes_i) or None,
            "last_mod_date": None,
        })
    return terms


# ---------------------------------------------------------------------------
# Merge logic
# ---------------------------------------------------------------------------

def merge_terms(tbx_terms: list[dict], gloss_terms: list[dict]) -> list[dict]:
    """Merge glossary terms into TBX terms. TBX is authoritative.
    Only add glossary entries whose source_term (case-insensitive) is not in TBX.
    """
    seen = {t["source_term"].lower() for t in tbx_terms}
    merged = list(tbx_terms)
    for gt in gloss_terms:
        key = gt["source_term"].lower()
        if key not in seen:
            seen.add(key)
            merged.append(gt)
    return merged


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    # 1. Parse all TBX files, grouped by language
    tbx_by_lang: dict[str, list[dict]] = {}
    for f in sorted(TBX_DIR.iterdir()):
        if not f.suffix.lower() == ".xlsx" or f.name.startswith("~$"):
            continue
        lang = extract_lang_from_tbx_filename(f.name)
        if not lang:
            print(f"  [WARN] Cannot extract lang from TBX: {f.name}", file=sys.stderr)
            continue
        print(f"  Parsing TBX: {f.name}  -> {lang}")
        terms = parse_tbx(f)
        tbx_by_lang[lang] = terms
        print(f"    {len(terms)} terms")

    # 2. Pick latest glossary per language
    gloss_files = pick_latest_glossary_per_lang(GLOSS_DIR)
    print()
    for lang, gf in sorted(gloss_files.items()):
        print(f"  Latest glossary for {lang}: {gf.name}")

    # 3. For each language, merge and output
    summary = {}
    all_langs = sorted(set(tbx_by_lang.keys()) | set(gloss_files.keys()))

    for lang in all_langs:
        tbx_terms = tbx_by_lang.get(lang, [])
        gloss_terms = []
        if lang in gloss_files:
            print(f"\n  Parsing glossary for {lang}: {gloss_files[lang].name}")
            gloss_terms = parse_glossary(gloss_files[lang])
            print(f"    {len(gloss_terms)} glossary terms")

        merged = merge_terms(tbx_terms, gloss_terms)

        # Assign term_id
        lang_upper = lang.upper()
        for idx, t in enumerate(merged, start=1):
            t["term_id"] = f"{lang_upper}_{idx:05d}"
            t["language"] = lang

        # Re-order keys to match schema
        ordered = []
        for t in merged:
            ordered.append({
                "term_id": t["term_id"],
                "language": t["language"],
                "source_term": t["source_term"],
                "target_term": t["target_term"],
                "part_of_speech": t.get("part_of_speech"),
                "definition": t.get("definition"),
                "context": t.get("context"),
                "status": t.get("status"),
                "notes": t.get("notes"),
                "last_mod_date": t.get("last_mod_date"),
            })

        output = {
            "language": lang,
            "total_terms": len(ordered),
            "terms": ordered,
        }

        out_path = OUT_DIR / f"{lang}.json"
        with open(out_path, "w", encoding="utf-8") as fp:
            json.dump(output, fp, ensure_ascii=False, indent=2)

        summary[lang] = {
            "total_terms": len(ordered),
            "from_tbx": len(tbx_terms),
            "from_glossary": len(ordered) - len(tbx_terms),
        }

        print(f"  -> {out_path.name}: {len(ordered)} terms "
              f"(TBX: {len(tbx_terms)}, Glossary added: {len(ordered) - len(tbx_terms)})")

    # 4. Write summary
    summary_output = {
        "total_languages": len(summary),
        "languages": {k: summary[k] for k in sorted(summary)},
    }
    summary_path = OUT_DIR / "_summary.json"
    with open(summary_path, "w", encoding="utf-8") as fp:
        json.dump(summary_output, fp, ensure_ascii=False, indent=2)
    print(f"\n  Summary written to {summary_path}")
    print(f"  Total languages: {len(summary)}")
    for lang, info in sorted(summary.items()):
        print(f"    {lang}: {info['total_terms']} terms")


if __name__ == "__main__":
    main()
